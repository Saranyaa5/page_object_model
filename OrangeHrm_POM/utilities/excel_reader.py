import openpyxl

def get_exceldata(path, sheet):
    excel_data = []
    workbook = openpyxl.load_workbook(path)
    sheet_data = workbook[sheet]
    total_rows = sheet_data.max_row
    total_columns = sheet_data.max_column

    for r in range(2, total_rows + 1): 
        row_data = []
        for c in range(1, total_columns + 1):
            cell_value = sheet_data.cell(row=r, column=c).value
            if cell_value is None:
                row_data.append("")
            else:
                row_data.append(cell_value)
        excel_data.append(row_data)

    return excel_data
